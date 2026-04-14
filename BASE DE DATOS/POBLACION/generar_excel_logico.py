import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

output_path = r"c:\Users\Personal\Documents\semestre 1-26\si1\organigrama\CRM-VISUAL\BASE DE DATOS\4_Modelo_Logico_Tablas_GRULAC.xlsx"

tablas = {
    # 1. MAESTROS E INDEPENDIENTES
    "ROLES": ["[PK] id_rol", "nombre_rol", "permisos_json"],
    "EMPLEADOS": ["[PK] id_empleado", "ci_documento", "nombre_completo", "telefono"],
    "PROVEEDORES": ["[PK] id_proveedor", "ci_nit", "razon_social", "estado_reputacion"],
    "CLIENTES": ["[PK] id_cliente", "nit_facturacion", "razon_social", "telefono"],
    "CATALOGO_ITEMS": ["[PK] id_item", "codigo_sku", "nombre_producto", "tipo_item", "unidad_medida", "stock_minimo"],

    # 2. SEGUIRDAD Y AUDITORÍA
    "USUARIOS": ["[PK] id_usuario", "[FK] id_rol", "[FK_1-a-1] id_empleado", "email_corporativo", "password_hash", "estado_acceso"],
    "BITACORA_AUDITORIA": ["[PK] id_log", "[FK] id_usuario", "accion_sql", "tabla_afectada", "fecha_hora", "old_data", "new_data"],

    # 3. RECEPCIÓN Y PRODUCCIÓN
    "RECEPCIONES_LECHE": ["[PK] id_recepcion", "[FK] id_proveedor", "[FK] id_laboratorista", "litros_recibidos", "temperatura_celsius", "acidez_ph", "estado_triage", "fecha_registro"],
    "RECETAS_BOM": ["[PK] id_receta", "[FK] id_item_resultado", "version_receta", "rendimiento_esperado_pct", "estado_activa"],
    "RECETA_INGREDIENTES": ["[PK] id_detalle_receta", "[FK] id_receta", "[FK] id_item_ingrediente", "cantidad_kgs_necesaria"],
    "ORDENES_PRODUCCION": ["[PK] id_orden", "[FK] id_jefe_produccion", "[FK] id_receta", "estado_lote", "litros_invertidos", "kilos_obtenidos_brutos", "fecha_inicio", "fecha_cierre"],
    "LOTE_PRODUCCION": ["[PK] id_lote", "[FK] id_orden", "codigo_lote", "fecha_produccion", "cantidad_producida", "unidad_medida", "estado", "observaciones"],
    "FICHAS_CALIDAD": ["[PK] id_ficha", "[FK] id_orden", "[FK] id_ingeniero_qa", "ph_final", "salinidad", "dictamen_qa", "observaciones_tecnicas", "fecha_evaluacion"],
    "MOVIMIENTOS_KARDEX": ["[PK] id_movimiento", "[FK] id_item", "[FK] id_orden_asociada", "tipo_operacion", "cantidad_kilos", "concepto_operacion", "fecha_hora"],

    # 4. COMERCIAL, VENTAS Y LOGÍSTICA
    "PEDIDOS_VENTAS": ["[PK] id_pedido", "[FK] id_cliente", "[FK] id_vendedor", "estado_reserva", "monto_total_bs", "fecha_reserva"],
    "FACTURA": ["[PK] id_factura", "[FK_1-a-1] id_pedido", "numero_factura", "fecha_emision", "subtotal", "impuesto", "total", "metodo_pago", "estado"],
    "DETALLE_PEDIDOS": ["[PK] id_detalle", "[FK] id_pedido", "[FK] id_item", "cantidad_pedida", "precio_unitario"],
    "DESPACHOS_LOGISTICOS": ["[PK] id_despacho", "[FK_1-a-1] id_pedido", "[FK] id_encargado_logistica", "placa_camion", "fecha_salida_ruta"],
    "DEVOLUCIONES_QA": ["[PK] id_devolucion", "[FK] id_despacho", "[FK] id_asesor_ventas", "motivo_rechazo", "kilos_devueltos", "requiere_reposicion_caliente", "fecha_registro"]
}

try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo_Logico_Maestro"
    
    current_col = 1
    
    # Formatos Visuales
    font_title = Font(bold=True, size=12, color="FFFFFF")
    fill_title = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    font_header = Font(bold=True)
    fill_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    for table_name, cols in tablas.items():
        # Fila 1: Título de la Tabla Combinada
        cell_title = ws.cell(row=1, column=current_col, value=table_name)
        cell_title.font = font_title
        cell_title.fill = fill_title
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(cols) - 1)
        cell_title.alignment = Alignment(horizontal="center")
        
        for i, col_full in enumerate(cols):
            # Limpiar y Detectar Llaves
            key_type = ""
            attr_name = col_full
            if str(col_full).startswith("[PK]"):
                key_type = "PK"
                attr_name = col_full.replace("[PK] ", "").strip()
            elif str(col_full).startswith("[FK]"):
                key_type = "FK"
                attr_name = col_full.replace("[FK] ", "").strip()
            elif str(col_full).startswith("[FK_1-a-1]"):
                key_type = "FK U"
                attr_name = col_full.replace("[FK_1-a-1] ", "").strip()
            
            # Fila 2: Tipo de Llave (PK, FK o vacío)
            cell_key = ws.cell(row=2, column=current_col + i, value=key_type)
            cell_key.font = font_header
            cell_key.fill = fill_header
            cell_key.alignment = Alignment(horizontal="center")
            
            # Fila 3: Atributos
            cell_attr = ws.cell(row=3, column=current_col + i, value=attr_name)
            cell_attr.font = font_header
            cell_attr.fill = fill_header
            
            # Auto-Ajustar Tamaño de Celda
            ws.column_dimensions[cell_attr.column_letter].width = max(len(attr_name) + 4, 15)
        
        # Saltar a la siguiente tabla dejando 1 columna en blanco de espaciado
        current_col += len(cols) + 1
        
    wb.save(output_path)
    print(f"Excel actualizado con exito en un solo sheet: {output_path}")

except Exception as e:
    import traceback
    traceback.print_exc()
